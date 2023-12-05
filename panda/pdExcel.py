#!/usr/bin/python3
import openpyxl

# 打开 Excel 文件
from openpyxl.utils import get_column_letter

from openpyxl.styles import Font, PatternFill

if __name__ == '__main__':
    workbook = openpyxl.load_workbook('a.xlsx')
    ws = workbook.active
    # 选择要操作的工作表
    worksheet = workbook['Sheet1']

    # 设置字体和背景颜色
    font = Font(name='黑体', size=15,bold=True)
    fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')

    for i in range(1, 5):
        cell = ws.cell(row=1, column=i)
        cell.font = font
        cell.fill = fill
        # 自适应调整列宽
    for column in worksheet.columns:
        max_length = 0
        column = get_column_letter(column[0].column)  # 获取列字母
        for cell in worksheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 10)  # 加 2 留有一定空间
        worksheet.column_dimensions[column].width = adjusted_width
        # 保存 Excel 文件
        workbook.save('a.xlsx')
    print("处理完成")